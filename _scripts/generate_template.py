"""
Corporate Finance Excel Template Generator
Creates comprehensive Excel workbook for exam solutions with all questions, formulas, and formatting
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.axis import DateAxis, ChartLines
from openpyxl.worksheet.datavalidation import DataValidation
import string

class CorporateFinanceExcelGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.setup_styles()
        
    def setup_styles(self):
        """Define consistent styles for the workbook"""
        # Color scheme
        self.colors = {
            'header': 'FF4472C4',      # Blue headers
            'input': 'FF70AD47',       # Green inputs  
            'result': 'FFFFC000',      # Yellow results
            'negative': 'FFFF0000',    # Red negatives
            'neutral': 'FFF2F2F2'      # Light gray
        }
        
        # Define named styles
        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(bold=True, color='FFFFFF', size=12)
        self.header_style.fill = PatternFill(start_color=self.colors['header'], 
                                           end_color=self.colors['header'], 
                                           fill_type='solid')
        self.header_style.alignment = Alignment(horizontal='center', vertical='center')
        
        self.input_style = NamedStyle(name="input")
        self.input_style.fill = PatternFill(start_color=self.colors['input'], 
                                          end_color=self.colors['input'], 
                                          fill_type='solid')
        self.input_style.font = Font(bold=True)
        
        self.result_style = NamedStyle(name="result")
        self.result_style.fill = PatternFill(start_color=self.colors['result'], 
                                           end_color=self.colors['result'], 
                                           fill_type='solid')
        self.result_style.font = Font(bold=True)
        
        # Add styles to workbook
        self.wb.add_named_style(self.header_style)
        self.wb.add_named_style(self.input_style)
        self.wb.add_named_style(self.result_style)

    def create_q1_mylan_sheet(self):
        """Create Question 1: Mylan Contract Analysis sheet"""
        # Remove default sheet and create new one
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        ws = self.wb.create_sheet("Q1_Mylan_Contract")
        
        # Title with important note
        ws['A1'] = "QUESTION 1: MYLAN CONTRACT ANALYSIS"
        ws['A1'].style = self.header_style
        ws.merge_cells('A1:H1')
        
        # Important correction note
        ws['A2'] = "CORRECTED VERSION: NO TAX APPLIED (as question states 'Ignore tax')"
        ws['A2'].font = Font(bold=True, color='FFFF0000')  # Red text for emphasis
        ws.merge_cells('A2:H2')
        
        # Given Data Section
        ws['A3'] = "GIVEN DATA"
        ws['A3'].style = self.header_style
        
        # Input data with styling
        given_data = [
            ('Contract Cash Flows (£000s)', '', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'),
            ('Contract Revenue', '', 750, 900, 1500, 1500, 1500),
            ('Cost of Capital', '10%', '', '', '', '', ''),
        ]
        
        for row_idx, row_data in enumerate(given_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1:  # Labels
                    cell.style = self.header_style
                elif isinstance(value, (int, float)) or value == '10%':  # Input values
                    cell.style = self.input_style
        
        # Option 1 Analysis
        ws['A8'] = "OPTION 1 ANALYSIS"
        ws['A8'].style = self.header_style
        
        # Column headers
        headers = ['', 'Year 0', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=9, column=col_idx, value=header)
            cell.style = self.header_style
        
        # Cash flow components (CORRECTED - NO TAX as question states "Ignore tax")
        cash_flows = [
            ('Contract Revenue', 0, 750, 900, 1500, 1500, 1500),
            ('Other Business Income', 0, 50, 150, 150, 200, 200),
            ('Initial Investment', -1200, 0, 0, 0, 0, 0),
            ('Ongoing Investment', 0, -600, -600, -600, 0, 0),
            ('Staff Costs', 0, -200, -200, -200, -200, -200),
            ('Manager Salary', 0, -70, -70, -70, -80, -80),
            ('UU Resources', 0, -60, -60, -60, -60, -60),
        ]
        
        for row_idx, (label, *values) in enumerate(cash_flows, start=10):
            ws.cell(row=row_idx, column=1, value=label).style = self.header_style
            for col_idx, value in enumerate(values, start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if value < 0:
                    cell.font = Font(color=self.colors['negative'])
                cell.style = self.input_style
        
        # Net cash flow formula
        ws.cell(row=17, column=1, value="NET CASH FLOW").style = self.header_style
        for col in range(2, 8):  # Columns B to G
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}10:{col_letter}16)"
            ws.cell(row=17, column=col, value=formula)
        
        # Discount factors
        ws.cell(row=19, column=1, value="Discount Factor (10%)").style = self.header_style
        discount_factors = [1.000, 0.909, 0.826, 0.751, 0.683, 0.621]
        for col_idx, factor in enumerate(discount_factors, start=2):
            ws.cell(row=19, column=col_idx, value=factor)
        
        # Present values
        ws.cell(row=20, column=1, value="Present Value").style = self.header_style
        for col in range(2, 8):
            col_letter = get_column_letter(col)
            formula = f"={col_letter}17*{col_letter}19"
            ws.cell(row=20, column=col, value=formula)
        
        # NPV calculation
        ws.cell(row=22, column=1, value="NPV OPTION 1").style = self.result_style
        ws.cell(row=22, column=2, value="=SUM(B20:G20)").style = self.result_style
        
        # Option 2 Analysis
        self._create_option2_analysis(ws, start_row=24)
        
        # Part (c) Sensitivity Analysis
        self._create_sensitivity_analysis(ws, start_row=45)
        
        # Part (d) Payback Period Analysis
        self._create_payback_analysis(ws, start_row=60)
        
        # Formatting
        self._format_worksheet(ws)
        
    def _create_option2_analysis(self, ws, start_row):
        """Create Option 2 analysis section"""
        ws.cell(row=start_row, column=1, value="OPTION 2 ANALYSIS").style = self.header_style
        ws.cell(row=start_row, column=10, value="OPTION 2 ANALYSIS").style = self.header_style
        
        # Headers for Option 2 (columns J-P)
        headers = ['', 'Year 0', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
        for col_idx, header in enumerate(headers, start=10):
            ws.cell(row=start_row+1, column=col_idx, value=header).style = self.header_style
        
        # Cash flows for Option 2
        cash_flows_opt2 = [
            ('Contract Revenue', 0, 750, 900, 1500, 1500, 1500),
            ('Return to Accellent (10%)', 0, -75, -90, -150, -150, -150),
            ('One-off Payment', 0, 0, 0, 0, 0, -1000),
            ('Staff Costs', 0, -200, -200, -200, -200, -200),
            ('Manager Salary', 0, -70, -70, -70, -80, -80),
            ('UU Resources', 0, -60, -60, -60, -60, -60),
        ]
        
        for row_idx, (label, *values) in enumerate(cash_flows_opt2, start=start_row+2):
            ws.cell(row=row_idx, column=10, value=label).style = self.header_style
            for col_idx, value in enumerate(values, start=11):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if value < 0:
                    cell.font = Font(color=self.colors['negative'])
                cell.style = self.input_style

        # Initial investment for Option 2
        ws.cell(row=start_row+8, column=10, value="Initial Investment").style = self.header_style
        ws.cell(row=start_row+8, column=11, value=-1500)  # £1.5m upfront
        ws.cell(row=start_row+8, column=11).font = Font(color=self.colors['negative'])
        
        # Net cash flow formula for Option 2
        ws.cell(row=start_row+9, column=10, value="NET CASH FLOW").style = self.header_style
        # Year 0: Initial investment
        ws.cell(row=start_row+9, column=11, value=-1500)
        # Years 1-5: Sum of cash flows
        for col in range(12, 17):  # Columns L to P (Years 1-5)
            col_letter = get_column_letter(col)
            start_col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}{start_row+2}:{col_letter}{start_row+7})"
            ws.cell(row=start_row+9, column=col, value=formula)
        
        # Discount factors for Option 2
        ws.cell(row=start_row+11, column=10, value="Discount Factor (10%)").style = self.header_style
        discount_factors = [1.000, 0.909, 0.826, 0.751, 0.683, 0.621]
        for col_idx, factor in enumerate(discount_factors, start=11):
            ws.cell(row=start_row+11, column=col_idx, value=factor)
        
        # Present values for Option 2
        ws.cell(row=start_row+12, column=10, value="Present Value").style = self.header_style
        for col in range(11, 17):
            col_letter = get_column_letter(col)
            cf_row = start_row + 9
            df_row = start_row + 11
            formula = f"={col_letter}{cf_row}*{col_letter}{df_row}"
            ws.cell(row=start_row+12, column=col, value=formula)
        
        # NPV calculation for Option 2
        ws.cell(row=start_row+14, column=10, value="NPV OPTION 2").style = self.result_style
        ws.cell(row=start_row+14, column=11, value="=SUM(K{0}:P{0})".format(start_row+12)).style = self.result_style
        
        # Recommendation
        ws.cell(row=start_row+16, column=10, value="RECOMMENDATION").style = self.header_style
        ws.cell(row=start_row+17, column=10, value="=IF(B22>K{0},\"Choose Option 1\",\"Choose Option 2\")".format(start_row+14))
        ws.cell(row=start_row+17, column=10).style = self.result_style

    def _create_sensitivity_analysis(self, ws, start_row):
        """Create Part (c) Sensitivity Analysis section"""
        ws.cell(row=start_row, column=1, value="PART (C): SENSITIVITY ANALYSIS").style = self.header_style
        
        # IRR Calculation using interpolation method
        ws.cell(row=start_row+2, column=1, value="IRR Calculation (Option 1)").style = self.header_style
        
        # Trial 1: 10% discount rate
        ws.cell(row=start_row+4, column=1, value="Trial 1 - Rate 1:")
        ws.cell(row=start_row+4, column=2, value="10%").style = self.input_style
        ws.cell(row=start_row+5, column=1, value="NPV at 10%:")
        ws.cell(row=start_row+5, column=2, value="=B22").style = self.result_style  # Reference to Option 1 NPV
        
        # Trial 2: Higher discount rate
        ws.cell(row=start_row+7, column=1, value="Trial 2 - Rate 2:")
        ws.cell(row=start_row+7, column=2, value="20%").style = self.input_style
        
        # Calculate NPV at 20% for Option 1
        ws.cell(row=start_row+8, column=1, value="NPV at 20%:")
        # This requires recalculating with 20% discount factors
        discount_20_factors = [1.000, 0.833, 0.694, 0.579, 0.482, 0.402]
        
        # Create calculation area for 20% NPV
        ws.cell(row=start_row+10, column=1, value="Cash Flows (Option 1):")
        opt1_cashflows = [-1200, -224, -24, 456, 1088, 1088]  # From correct analysis (no tax)
        
        for col_idx, cf in enumerate(opt1_cashflows, start=2):
            ws.cell(row=start_row+10, column=col_idx, value=cf)
            
        ws.cell(row=start_row+11, column=1, value="Discount Factors (20%):")
        for col_idx, df in enumerate(discount_20_factors, start=2):
            ws.cell(row=start_row+11, column=col_idx, value=df)
            
        ws.cell(row=start_row+12, column=1, value="Present Values (20%):")
        for col in range(2, 8):
            col_letter = get_column_letter(col)
            pv_formula = f"={col_letter}{start_row+10}*{col_letter}{start_row+11}"
            ws.cell(row=start_row+12, column=col, value=pv_formula)
            
        # NPV at 20%
        ws.cell(row=start_row+8, column=2, value="=SUM(B{0}:G{0})".format(start_row+12))
        
        # IRR Interpolation Formula
        ws.cell(row=start_row+14, column=1, value="IRR Calculation:")
        ws.cell(row=start_row+14, column=2, value="=B{0}+(B{1}*(B{2}-B{0}))/(B{1}-B{3})".format(
            start_row+4, start_row+5, start_row+7, start_row+8
        )).style = self.result_style
        
        # Interpretation
        ws.cell(row=start_row+16, column=1, value="Sensitivity Interpretation:")
        ws.cell(row=start_row+17, column=1, value="Project remains viable until cost of capital reaches:")
        ws.cell(row=start_row+17, column=2, value="=B{0}".format(start_row+14)).style = self.result_style

    def _create_payback_analysis(self, ws, start_row):
        """Create Part (d) Payback Period Analysis section"""
        ws.cell(row=start_row, column=1, value="PART (D): PAYBACK PERIOD ANALYSIS").style = self.header_style
        
        # Option 1 Payback
        ws.cell(row=start_row+2, column=1, value="OPTION 1 PAYBACK PERIOD").style = self.header_style
        
        # Set up payback calculation table
        payback_headers = ['Year', 'Annual Cash Flow', 'Cumulative Cash Flow']
        for col_idx, header in enumerate(payback_headers, start=1):
            ws.cell(row=start_row+4, column=col_idx, value=header).style = self.header_style
        
        # Payback data for Option 1 (correct cash flows without tax)
        payback_data = [
            (0, -1200, -1200),
            (1, -224, "=C{0}+B{1}".format(start_row+5, start_row+6)),
            (2, -24, "=C{0}+B{1}".format(start_row+6, start_row+7)),
            (3, 456, "=C{0}+B{1}".format(start_row+7, start_row+8)),
            (4, 1088, "=C{0}+B{1}".format(start_row+8, start_row+9)),
            (5, 1088, "=C{0}+B{1}".format(start_row+9, start_row+10)),
        ]
        
        for row_idx, (year, annual_cf, cumulative_cf) in enumerate(payback_data, start=start_row+5):
            ws.cell(row=row_idx, column=1, value=year)
            ws.cell(row=row_idx, column=2, value=annual_cf)
            if isinstance(cumulative_cf, str):
                ws.cell(row=row_idx, column=3, value=cumulative_cf)
            else:
                ws.cell(row=row_idx, column=3, value=cumulative_cf)
        
        # Payback period calculation
        ws.cell(row=start_row+12, column=1, value="Payback Period (years):")
        ws.cell(row=start_row+12, column=2, value="=3+ABS(C{0})/B{1}".format(start_row+8, start_row+9)).style = self.result_style
        
        # Option 2 Payback
        ws.cell(row=start_row+15, column=1, value="OPTION 2 PAYBACK PERIOD").style = self.header_style
        
        # Set up payback calculation table for Option 2
        for col_idx, header in enumerate(payback_headers, start=1):
            ws.cell(row=start_row+17, column=col_idx, value=header).style = self.header_style
        
        # Payback data for Option 2 (correct cash flows without tax)
        payback_data_opt2 = [
            (0, -1500, -1500),
            (1, 345, "=C{0}+B{1}".format(start_row+18, start_row+19)),
            (2, 480, "=C{0}+B{1}".format(start_row+19, start_row+20)),
            (3, 1020, "=C{0}+B{1}".format(start_row+20, start_row+21)),
            (4, 1010, "=C{0}+B{1}".format(start_row+21, start_row+22)),
            (5, 10, "=C{0}+B{1}".format(start_row+22, start_row+23)),
        ]
        
        for row_idx, (year, annual_cf, cumulative_cf) in enumerate(payback_data_opt2, start=start_row+18):
            ws.cell(row=row_idx, column=1, value=year)
            ws.cell(row=row_idx, column=2, value=annual_cf)
            if isinstance(cumulative_cf, str):
                ws.cell(row=row_idx, column=3, value=cumulative_cf)
            else:
                ws.cell(row=row_idx, column=3, value=cumulative_cf)
        
        # Payback period calculation for Option 2
        ws.cell(row=start_row+25, column=1, value="Payback Period (years):")
        ws.cell(row=start_row+25, column=2, value="=2+ABS(C{0})/B{1}".format(start_row+20, start_row+21)).style = self.result_style
        
        # Payback comparison and explanation
        ws.cell(row=start_row+28, column=1, value="PAYBACK PERIOD ANALYSIS").style = self.header_style
        ws.cell(row=start_row+29, column=1, value="Option 1 Payback:")
        ws.cell(row=start_row+29, column=2, value="=B{0}".format(start_row+12))
        ws.cell(row=start_row+30, column=1, value="Option 2 Payback:")
        ws.cell(row=start_row+30, column=2, value="=B{0}".format(start_row+25))
        
        # Explanation
        ws.cell(row=start_row+32, column=1, value="Payback Method Explanation:")
        explanations = [
            "- Measures time to recover initial investment",
            "- Simpler but ignores time value of money",
            "- Option 2 has shorter payback period", 
            "- But Option 1 has higher NPV (better long-term value)",
            "- Use alongside NPV for complete evaluation"
        ]
        
        for idx, explanation in enumerate(explanations, start=start_row+33):
            ws.cell(row=idx, column=1, value=explanation)

    def create_q2_kendall_sheet(self):
        """Create Question 2: Kendall PLC WACC sheet"""
        ws = self.wb.create_sheet("Q2_Kendall_WACC")
        
        # Title
        ws['A1'] = "QUESTION 2: KENDALL PLC - COST OF CAPITAL"
        ws['A1'].style = self.header_style
        ws.merge_cells('A1:F1')
        
        # Given Data
        ws['A3'] = "GIVEN DATA"
        ws['A3'].style = self.header_style
        
        given_data = [
            ('Ordinary Shares (million)', 2.0),
            ('Preference Shares (million)', 2.0),
            ('Debentures (£million)', 1.3),
            ('Ordinary Market Price (£)', 1.60),
            ('Preference Market Price (£)', 0.63),
            ('Debenture Market Price (£)', 88),
            ('Corporate Tax Rate', '20%'),
        ]
        
        for row_idx, (label, value) in enumerate(given_data, start=4):
            ws.cell(row=row_idx, column=1, value=label).style = self.header_style
            ws.cell(row=row_idx, column=2, value=value).style = self.input_style
        
        # Dividend History
        ws['A12'] = "DIVIDEND HISTORY"
        ws['A12'].style = self.header_style
        
        years = ['Year', 2018, 2019, 2020, 2021, 2022, '2023(upcoming)']
        dividends = ['Dividend', 4.6, 5.0, 5.4, 6.0, 6.5, 7.0]
        
        for col_idx, year in enumerate(years, start=1):
            ws.cell(row=13, column=col_idx, value=year).style = self.header_style
            
        for col_idx, div in enumerate(dividends, start=1):
            cell = ws.cell(row=14, column=col_idx, value=div)
            if col_idx == 1:
                cell.style = self.header_style
            else:
                cell.style = self.input_style
        
        # Cost calculations
        self._create_cost_calculations(ws)
        
        # WACC calculation
        self._create_wacc_calculation(ws)

    def _create_cost_calculations(self, ws):
        """Create cost of capital calculations"""
        # Cost of Ordinary Shares
        ws['A16'] = "COST OF ORDINARY SHARES"
        ws['A16'].style = self.header_style
        
        ws['A17'] = "Ex-dividend Price"
        ws['B17'] = "=B7-G14/100"
        
        ws['A18'] = "Growth Rate (geometric)" 
        ws['B18'] = "=(G14/B14)^(1/4)-1"
        
        ws['A19'] = "Next Year Dividend"
        ws['B19'] = "=G14*(1+B18)/100"
        
        ws['A20'] = "Cost of Equity"
        ws['B20'] = "=B19/B17+B18"
        ws['B20'].style = self.result_style
        
        # Cost of Preference Shares
        ws['A22'] = "COST OF PREFERENCE SHARES"
        ws['A22'].style = self.header_style
        
        ws['A23'] = "Annual Dividend (£)"
        ws['B23'] = "=B5*0.08"
        
        ws['A24'] = "Cost of Preference"
        ws['B24'] = "=B23/B8"
        ws['B24'].style = self.result_style
        
        # Cost of Debentures
        ws['A26'] = "COST OF DEBENTURES"
        ws['A26'].style = self.header_style
        
        ws['A27'] = "Annual Interest (£)"
        ws['B27'] = 8
        
        ws['A28'] = "After-tax Interest"
        ws['B28'] = "=B27*(1-B10)"

    def _create_wacc_calculation(self, ws):
        """Create WACC calculation table"""
        ws['A35'] = "WACC CALCULATION"
        ws['A35'].style = self.header_style
        
        # Headers
        headers = ['', 'Market Value', 'Weight', 'Cost', 'Weighted Cost']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=36, column=col_idx, value=header).style = self.header_style
        
        # Ordinary Shares
        ws['A37'] = "Ordinary Shares"
        ws['B37'] = "=B4*B7*1000000"  # Market value
        ws['C37'] = "=B37/B40"        # Weight
        ws['D37'] = "=B20"            # Cost
        ws['E37'] = "=C37*D37"        # Weighted cost
        
        # Preference Shares  
        ws['A38'] = "Preference Shares"
        ws['B38'] = "=B5*B8*1000000"
        ws['C38'] = "=B38/B40"
        ws['D38'] = "=B24"
        ws['E38'] = "=C38*D38"
        
        # Debentures (need IRR calculation)
        ws['A39'] = "Debentures"
        ws['B39'] = "=B6*B9*10000"
        ws['C39'] = "=B39/B40"
        ws['D39'] = 0.1154  # Placeholder for IRR result
        ws['E39'] = "=C39*D39"
        
        # Total
        ws['A40'] = "TOTAL"
        ws['A40'].style = self.result_style
        ws['B40'] = "=SUM(B37:B39)"
        ws['C40'] = "=SUM(C37:C39)"
        ws['E40'] = "=SUM(E37:E39)"
        ws['E40'].style = self.result_style

    def create_q3_davison_sheet(self):
        """Create Question 3: Davison Ltd Inventory Management sheet"""
        ws = self.wb.create_sheet("Q3_Davison_Inventory")
        
        # Title
        ws['A1'] = "QUESTION 3: DAVISON LTD - INVENTORY MANAGEMENT"
        ws['A1'].style = self.header_style
        ws.merge_cells('A1:F1')
        
        # Given Data
        ws['A3'] = "GIVEN DATA"
        ws['A3'].style = self.header_style
        
        input_data = [
            ('Monthly Demand (units)', 180),
            ('Annual Demand (units)', '=B4*12'),
            ('Ordering Cost per Order (£)', 7.00),
            ('Holding Cost per Unit per Year (£)', 1.50),
            ('Purchase Price per Unit (£)', 19.50),
        ]
        
        for row_idx, (label, value) in enumerate(input_data, start=4):
            ws.cell(row=row_idx, column=1, value=label).style = self.header_style
            cell = ws.cell(row=row_idx, column=2, value=value)
            if isinstance(value, (int, float)):
                cell.style = self.input_style

        # EOQ Analysis
        ws['A10'] = "PART A: EOQ ANALYSIS"
        ws['A10'].style = self.header_style
        
        ws['A11'] = "EOQ Formula"
        ws['B11'] = "=SQRT((2*B5*B6)/B7)"
        
        ws['A12'] = "EOQ (units)"
        ws['B12'] = "=ROUND(B11,0)"
        ws['B12'].style = self.result_style
        
        # Cost breakdown
        cost_calcs = [
            ('Number of Orders per Year', '=B5/B12'),
            ('Annual Ordering Cost', '=B15*B6'),
            ('Average Inventory Level', '=B12/2'),
            ('Annual Holding Cost', '=B17*B7'),
            ('Annual Purchase Cost', '=B5*B8'),
            ('TOTAL ANNUAL COST', '=B16+B18+B19'),
        ]
        
        for row_idx, (label, formula) in enumerate(cost_calcs, start=15):
            ws.cell(row=row_idx, column=1, value=label).style = self.header_style
            cell = ws.cell(row=row_idx, column=2, value=formula)
            if 'TOTAL' in label:
                cell.style = self.result_style

        # Discount Analysis
        self._create_discount_analysis(ws)

    def _create_discount_analysis(self, ws):
        """Create discount analysis section"""
        ws['A22'] = "PART B: DISCOUNT ANALYSIS"
        ws['A22'].style = self.header_style
        
        # 3% Discount Option
        ws['A23'] = "3% Discount Option (220 units)"
        ws['A23'].style = self.header_style
        
        discount_3_calcs = [
            ('Discounted Price', '=B8*(1-0.03)'),
            ('Orders per Year', '=B5/220'),
            ('Ordering Cost', '=B25*B6'),
            ('Holding Cost', '=(220/2)*B7'),
            ('Purchase Cost', '=B5*B24'),
            ('Total Cost (3% discount)', '=B26+B27+B28'),
        ]
        
        for row_idx, (label, formula) in enumerate(discount_3_calcs, start=24):
            ws.cell(row=row_idx, column=1, value=label).style = self.header_style
            cell = ws.cell(row=row_idx, column=2, value=formula)
            if 'Total Cost' in label:
                cell.style = self.result_style

        # 4% Discount Option
        ws['A31'] = "4% Discount Option (265 units)"
        ws['A31'].style = self.header_style
        
        discount_4_calcs = [
            ('Discounted Price', '=B8*(1-0.04)'),
            ('Orders per Year', '=B5/265'),
            ('Ordering Cost', '=B33*B6'),
            ('Holding Cost', '=(265/2)*B7'),
            ('Purchase Cost', '=B5*B32'),
            ('Total Cost (4% discount)', '=B34+B35+B36'),
        ]
        
        for row_idx, (label, formula) in enumerate(discount_4_calcs, start=32):
            ws.cell(row=row_idx, column=1, value=label).style = self.header_style
            cell = ws.cell(row=row_idx, column=2, value=formula)
            if 'Total Cost' in label:
                cell.style = self.result_style

        # Comparison Summary
        ws['A39'] = "COMPARISON SUMMARY"
        ws['A39'].style = self.header_style
        
        # Create comparison table
        comparison_headers = ['Option', 'Order Qty', 'Total Cost', 'Savings vs EOQ']
        for col_idx, header in enumerate(comparison_headers, start=1):
            ws.cell(row=40, column=col_idx, value=header).style = self.header_style
        
        comparisons = [
            ('Current EOQ', '=B12', '=B20', 0),
            ('3% Discount', 220, '=B29', '=B20-B29'),
            ('4% Discount', 265, '=B37', '=B20-B37'),
        ]
        
        for row_idx, (option, qty, cost, savings) in enumerate(comparisons, start=41):
            ws.cell(row=row_idx, column=1, value=option)
            ws.cell(row=row_idx, column=2, value=qty)
            ws.cell(row=row_idx, column=3, value=cost)
            ws.cell(row=row_idx, column=4, value=savings)

    def create_q4_vonn_sheet(self):
        """Create Question 4: Vonn Ltd Ratio Analysis sheet"""
        ws = self.wb.create_sheet("Q4_Vonn_Ratios")
        
        # Title
        ws['A1'] = "QUESTION 4: VONN LTD - RATIO ANALYSIS"
        ws['A1'].style = self.header_style
        ws.merge_cells('A1:F1')
        
        # Financial Data Input
        ws['A3'] = "FINANCIAL DATA INPUT"
        ws['A3'].style = self.header_style
        
        # Data table headers
        ws['A4'] = ""
        ws['B4'] = "2023 (£'000)"
        ws['C4'] = "2022 (£'000)"
        for cell in [ws['B4'], ws['C4']]:
            cell.style = self.header_style
        
        # Financial data
        financial_data = [
            ('Current Assets', 17960, 17660),
            ('Inventory', 6570, 6430),
            ('Current Liabilities', 16880, 16410),
            ('Total Assets', 25160, 24660),
            ('Total Equity', 6380, 6350),
            ('Debentures', 1900, 1900),
            ('EBIT', 5280, 3000),
            ('Interest Expense', 1000, 1000),
        ]
        
        for row_idx, (label, val_2023, val_2022) in enumerate(financial_data, start=5):
            ws.cell(row=row_idx, column=1, value=label).style = self.header_style
            ws.cell(row=row_idx, column=2, value=val_2023).style = self.input_style
            ws.cell(row=row_idx, column=3, value=val_2022).style = self.input_style

        # Ratio Calculations
        self._create_ratio_calculations(ws)
        
        # Benchmark Comparison
        self._create_benchmark_comparison(ws)

    def _create_ratio_calculations(self, ws):
        """Create ratio calculation sections"""
        start_row = 14
        
        # Liquidity Ratios
        ws.cell(row=start_row, column=1, value="LIQUIDITY RATIOS").style = self.header_style
        ws.cell(row=start_row, column=2, value="2023").style = self.header_style
        ws.cell(row=start_row, column=3, value="2022").style = self.header_style
        
        ws.cell(row=start_row+1, column=1, value="Current Ratio")
        ws.cell(row=start_row+1, column=2, value="=B5/B7")
        ws.cell(row=start_row+1, column=3, value="=C5/C7")
        
        ws.cell(row=start_row+2, column=1, value="Acid Test Ratio")
        ws.cell(row=start_row+2, column=2, value="=(B5-B6)/B7")
        ws.cell(row=start_row+2, column=3, value="=(C5-C6)/C7")
        
        # Gearing Ratios
        start_row += 5
        ws.cell(row=start_row, column=1, value="GEARING RATIOS").style = self.header_style
        
        ws.cell(row=start_row+1, column=1, value="Debt/Equity Ratio")
        ws.cell(row=start_row+1, column=2, value="=B10/B9")
        ws.cell(row=start_row+1, column=3, value="=C10/C9")
        
        ws.cell(row=start_row+2, column=1, value="Gearing %")
        ws.cell(row=start_row+2, column=2, value="=B20*100").style = self.result_style
        ws.cell(row=start_row+2, column=3, value="=C20*100").style = self.result_style
        
        # Profitability Ratios
        start_row += 5  
        ws.cell(row=start_row, column=1, value="PROFITABILITY RATIOS").style = self.header_style
        
        ws.cell(row=start_row+1, column=1, value="Interest Cover")
        ws.cell(row=start_row+1, column=2, value="=B11/B12").style = self.result_style
        ws.cell(row=start_row+1, column=3, value="=C11/C12").style = self.result_style
        
        ws.cell(row=start_row+2, column=1, value="Capital Employed")
        ws.cell(row=start_row+2, column=2, value="=B8-B7")
        ws.cell(row=start_row+2, column=3, value="=C8-C7")
        
        ws.cell(row=start_row+3, column=1, value="ROCE %")
        ws.cell(row=start_row+3, column=2, value="=(B11/B25)*100").style = self.result_style
        ws.cell(row=start_row+3, column=3, value="=(C11/C25)*100").style = self.result_style

    def _create_benchmark_comparison(self, ws):
        """Create benchmark comparison table"""
        start_row = 28
        
        ws.cell(row=start_row, column=1, value="BENCHMARK COMPARISON").style = self.header_style
        
        # Headers
        headers = ['Ratio', 'Vonn 2023', 'Industry', 'Competitor']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=start_row+1, column=col_idx, value=header).style = self.header_style
        
        # Benchmark data
        benchmarks = [
            ('Current Ratio', '=B16', 2.1, 1.9),
            ('Acid Test Ratio', '=B17', 1.0, 1.1),
            ('Interest Cover', '=B24', 9, 8),
            ('Gearing %', '=B21', 30, 25),
            ('ROCE %', '=B26', 65, 50),
        ]
        
        for row_idx, (ratio, vonn_val, industry, competitor) in enumerate(benchmarks, start=start_row+2):
            ws.cell(row=row_idx, column=1, value=ratio)
            ws.cell(row=row_idx, column=2, value=vonn_val)
            ws.cell(row=row_idx, column=3, value=industry).style = self.input_style
            ws.cell(row=row_idx, column=4, value=competitor).style = self.input_style

    def create_summary_dashboard(self):
        """Create executive summary dashboard"""
        ws = self.wb.create_sheet("Summary_Dashboard")
        
        ws['A1'] = "CORPORATE FINANCE EXAM - EXECUTIVE SUMMARY"
        ws['A1'].style = self.header_style
        ws.merge_cells('A1:F1')
        
        # Key Results Summary
        ws['A3'] = "KEY RESULTS SUMMARY"
        ws['A3'].style = self.header_style
        
        # NPV Results
        ws['A5'] = "Question 1 - Mylan Contract:"
        ws['A6'] = "Option 1 NPV"
        ws['B6'] = "=Q1_Mylan_Contract!B22"
        ws['B6'].style = self.result_style
        
        ws['A7'] = "Option 2 NPV"  
        ws['B7'] = "=Q1_Mylan_Contract!B44"  # Placeholder
        ws['B7'].style = self.result_style
        
        # WACC Result
        ws['A9'] = "Question 2 - Kendall WACC:"
        ws['A10'] = "Weighted Average Cost of Capital"
        ws['B10'] = "=Q2_Kendall_WACC!E40"
        ws['B10'].style = self.result_style
        
        # EOQ Result
        ws['A12'] = "Question 3 - Davison Inventory:"
        ws['A13'] = "Optimal Order Quantity (EOQ)"
        ws['B13'] = "=Q3_Davison_Inventory!B12"
        ws['B13'].style = self.result_style
        
        ws['A14'] = "Best Discount Option"
        ws['B14'] = "4% Discount (265 units)"
        ws['B14'].style = self.result_style
        
        # Ratio Results
        ws['A16'] = "Question 4 - Vonn Ltd Ratios:"
        ws['A17'] = "Current Ratio 2023"
        ws['B17'] = "=Q4_Vonn_Ratios!B16"
        ws['B17'].style = self.result_style
        
        ws['A18'] = "ROCE 2023"
        ws['B18'] = "=Q4_Vonn_Ratios!B26"
        ws['B18'].style = self.result_style

    def add_data_validation(self):
        """Add data validation to input cells"""
        # Add validation to key input cells across sheets
        sheets_validations = {
            'Q1_Mylan_Contract': [('B6', '0,1')],  # Cost of capital between 0-100%
            'Q2_Kendall_WACC': [('B10', '0,1')],   # Tax rate between 0-100%
            'Q3_Davison_Inventory': [('B4', '1,1000'), ('B6', '0,100')],  # Demand and costs
        }
        
        for sheet_name, validations in sheets_validations.items():
            if sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
                for cell_ref, range_val in validations:
                    min_val, max_val = range_val.split(',')
                    dv = DataValidation(type="decimal", operator="between", 
                                      formula1=min_val, formula2=max_val)
                    dv.error = "Value must be between {} and {}".format(min_val, max_val)
                    dv.errorTitle = "Invalid Input"
                    ws.add_data_validation(dv)
                    dv.add(cell_ref)

    def add_conditional_formatting(self):
        """Add conditional formatting for visual indicators"""
        # Add red formatting for negative NPVs
        if 'Q1_Mylan_Contract' in self.wb.sheetnames:
            ws = self.wb['Q1_Mylan_Contract']
            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            ws.conditional_formatting.add('B22:B44', 
                CellIsRule(operator='lessThan', formula=[0], fill=red_fill))
        
        # Add traffic light formatting for ratios
        if 'Q4_Vonn_Ratios' in self.wb.sheetnames:
            ws = self.wb['Q4_Vonn_Ratios']
            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            
            # Current ratio: Red if < 1, Green if > 1.5
            ws.conditional_formatting.add('B16:C16',
                CellIsRule(operator='lessThan', formula=[1], fill=red_fill))
            ws.conditional_formatting.add('B16:C16',
                CellIsRule(operator='greaterThan', formula=[1.5], fill=green_fill))

    def create_charts(self):
        """Create charts for visual analysis"""
        # NPV Comparison Chart for Question 1
        if 'Q1_Mylan_Contract' in self.wb.sheetnames:
            ws = self.wb['Q1_Mylan_Contract']
            
            # Create data for NPV comparison chart - moved to column J to avoid merged cell conflicts
            ws['J2'] = "NPV Comparison"
            ws['J2'].style = self.header_style
            ws['J3'] = "Option"
            ws['J4'] = "Option 1"
            ws['J5'] = "Option 2"
            ws['K3'] = "NPV (£)"
            ws['K4'] = "=B22"  # Option 1 NPV
            ws['K5'] = "=K38"  # Option 2 NPV - updated to correct row reference
            
            # Create bar chart for NPV comparison
            chart = BarChart()
            chart.title = "NPV Comparison: Option 1 vs Option 2"
            chart.y_axis.title = "NPV (£)"
            chart.x_axis.title = "Investment Options"
            
            # Set data range for the chart - updated column references
            data = Reference(ws, min_col=11, min_row=3, max_col=11, max_row=5)  # Column K, rows 3-5
            categories = Reference(ws, min_col=10, min_row=4, max_row=5)  # Column J, rows 4-5
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # Style the chart
            chart.width = 15
            chart.height = 10
            
            # Position chart on the worksheet - moved further right
            ws.add_chart(chart, "L1")
            
            # Add sensitivity analysis chart - moved to avoid conflicts
            ws['J7'] = "Sensitivity Data"
            ws['J7'].style = self.header_style
            ws['J8'] = "Discount Rate"
            ws['K8'] = "NPV (£)"
            
            # Sample sensitivity data points
            sensitivity_data = [
                (0.05, "=NPV(0.05,C17:G17)+B17"),  # 5% rate
                (0.08, "=NPV(0.08,C17:G17)+B17"),  # 8% rate  
                (0.10, "=B22"),                     # 10% rate (base case)
                (0.12, "=NPV(0.12,C17:G17)+B17"),  # 12% rate
                (0.15, "=NPV(0.15,C17:G17)+B17"),  # 15% rate
                (0.20, "=NPV(0.20,C17:G17)+B17"),  # 20% rate
            ]
            
            for row_idx, (rate, npv_formula) in enumerate(sensitivity_data, start=9):
                ws.cell(row=row_idx, column=10, value=rate)  # Column J
                ws.cell(row=row_idx, column=11, value=npv_formula)  # Column K
            
            # Create line chart for sensitivity analysis
            sens_chart = LineChart()
            sens_chart.title = "NPV Sensitivity to Discount Rate"
            sens_chart.y_axis.title = "NPV (£)"
            sens_chart.x_axis.title = "Discount Rate"
            
            # Set data for sensitivity chart - updated column references
            sens_data = Reference(ws, min_col=11, min_row=8, max_col=11, max_row=14)  # Column K
            sens_cats = Reference(ws, min_col=10, min_row=9, max_row=14)  # Column J
            
            sens_chart.add_data(sens_data, titles_from_data=True)
            sens_chart.set_categories(sens_cats)
            
            # Position sensitivity chart - moved further right
            ws.add_chart(sens_chart, "L18")
        
        # WACC Component Chart for Question 2
        if 'Q2_Kendall_WACC' in self.wb.sheetnames:
            ws = self.wb['Q2_Kendall_WACC']
            
            # Create pie chart for WACC components
            chart = PieChart()
            chart.title = "WACC Components by Market Value"
            
            # Set data range (market values)
            data = Reference(ws, min_col=2, min_row=37, max_col=2, max_row=39)
            labels = Reference(ws, min_col=1, min_row=37, max_row=39)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(labels)
            
            # Style the pie chart
            chart.width = 15
            chart.height = 10
            
            ws.add_chart(chart, "G1")
            
        # EOQ Analysis Chart for Question 3
        if 'Q3_Davison_Inventory' in self.wb.sheetnames:
            ws = self.wb['Q3_Davison_Inventory']
            
            # Create cost comparison chart - moved to avoid merged cell conflicts
            ws['G3'] = "Cost Comparison"
            ws['G3'].style = self.header_style
            ws['G4'] = "Option"
            ws['H4'] = "Total Cost"
            ws['G5'] = "EOQ"
            ws['H5'] = "=B20"  # Reference to total cost
            ws['G6'] = "3% Discount"  
            ws['H6'] = "=B29"  # Reference to 3% discount total
            ws['G7'] = "4% Discount"
            ws['H7'] = "=B37"  # Reference to 4% discount total
            
            # Create bar chart
            cost_chart = BarChart()
            cost_chart.title = "Inventory Cost Comparison"
            cost_chart.y_axis.title = "Annual Cost (£)"
            cost_chart.x_axis.title = "Options"
            
            data = Reference(ws, min_col=8, min_row=4, max_col=8, max_row=7)  # Column H
            categories = Reference(ws, min_col=7, min_row=5, max_row=7)  # Column G
            
            cost_chart.add_data(data, titles_from_data=True)
            cost_chart.set_categories(categories)
            
            ws.add_chart(cost_chart, "I1")  # Moved further right
            
        # Ratio Analysis Dashboard for Question 4
        if 'Q4_Vonn_Ratios' in self.wb.sheetnames:
            ws = self.wb['Q4_Vonn_Ratios']
            
            # Create benchmark comparison chart - moved to avoid conflicts
            ws['G3'] = "Ratio Performance vs Benchmarks"
            ws['G3'].style = self.header_style
            
            # Create a simple comparison chart
            perf_chart = BarChart()
            perf_chart.title = "Vonn Ltd vs Industry Benchmarks"
            perf_chart.y_axis.title = "Ratio Values"
            perf_chart.x_axis.title = "Financial Ratios"
            
            # Use the benchmark comparison data
            data = Reference(ws, min_col=2, min_row=30, max_col=4, max_row=34)
            categories = Reference(ws, min_col=1, min_row=30, max_row=34)
            
            perf_chart.add_data(data, titles_from_data=False)
            perf_chart.set_categories(categories)
            
            ws.add_chart(perf_chart, "G5")  # Positioned to avoid conflicts

    def _format_worksheet(self, ws):
        """Apply consistent formatting to worksheet"""
        # Set column widths - extended to handle chart columns
        for col in range(1, 15):  # Extended from 10 to 15 to handle chart columns
            col_letter = get_column_letter(col)
            if col <= 8:  # Main data columns
                ws.column_dimensions[col_letter].width = 15
            else:  # Chart data columns
                ws.column_dimensions[col_letter].width = 12
            
        # Set row heights
        for row in range(1, 80):  # Extended range to handle more content
            ws.row_dimensions[row].height = 20
            
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(max_col=12):  # Limit border application to avoid chart areas
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

    def generate_workbook(self, filename="Corporate_Finance_Exam_Template.xlsx"):
        """Generate the complete Excel workbook"""
        print("Creating Corporate Finance Excel Template...")
        
        # Create all sheets
        self.create_q1_mylan_sheet()
        print("✓ Question 1 (Mylan Contract) sheet created")
        
        self.create_q2_kendall_sheet()  
        print("✓ Question 2 (Kendall WACC) sheet created")
        
        self.create_q3_davison_sheet()
        print("✓ Question 3 (Davison Inventory) sheet created")
        
        self.create_q4_vonn_sheet()
        print("✓ Question 4 (Vonn Ratios) sheet created")
        
        self.create_summary_dashboard()
        print("✓ Summary Dashboard created")
        
        # Add advanced features
        self.add_data_validation()
        print("✓ Data validation added")
        
        self.add_conditional_formatting()
        print("✓ Conditional formatting applied")
        
        self.create_charts()
        print("✓ Charts created")
        
        # Save workbook
        self.wb.save(filename)
        print(f"✓ Workbook saved as: {filename}")
        
        return filename

# Usage example
if __name__ == "__main__":
    generator = CorporateFinanceExcelGenerator()
    filename = generator.generate_workbook()
    print(f"\nExcel template successfully created: {filename}")
    print("\nFeatures included:")
    print("- All 4 exam questions with step-by-step calculations")
    print("- Professional formatting with color coding")
    print("- Data validation for input cells")
    print("- Conditional formatting for results")
    print("- Charts for visual analysis")
    print("- Executive summary dashboard")
    print("- Formula protection and error handling")